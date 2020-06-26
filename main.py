import tkinter
import tkinter.filedialog as filedialog
from openpyxl import load_workbook

# define global variables
wbFirst = None
wsFirst = None

wbSecond = None
wsSecond = None

def selectFilesFunction():
    GUI = tkinter.Tk()
    GUI.geometry('640x640')

    def setFirstExcel():
        global wbFirst
        global wsFirst
        file_path = filedialog.askopenfilename()
        if(file_path):
            wbFirst = load_workbook(file_path)
            wsFirst = wbFirst.active
        

    def setSecondExcel():
        global wbSecond
        global wsSecond
        file_path = filedialog.askopenfilename()
        if(file_path):
            wbSecond = load_workbook(file_path)
            wsSecond = wbSecond.active

    B1 = tkinter.Button(GUI, text = "Excel1", command = setFirstExcel)
    B2 = tkinter.Button(GUI, text = "Excel2", command = setSecondExcel)
    B3 = tkinter.Button(GUI, text = "Compare", command = mainCompareFunction)

    B1.pack()
    B2.pack()
    B3.pack()

    GUI.mainloop()

def mainCompareFunction():
    global wbFirst
    global wbSecond
    global wsFirst
    global wsSecond
    class RowObject:

        def __init__(self, firstName, lastName, email):
            self.firstName = str(firstName).lower()
            self.lastName = str(lastName).lower()
            self.email = str(email).lower()
        
        def __repr__(self):
            return "%s %s %s" % (self.firstName, self.lastName, self.email)

        def __str__(self):
            return "%s %s %s" % (self.firstName, self.lastName, self.email)

    firstExcelArray = []
    firstExcelFirstNameCol = 0
    firstExcelLastNameCol = 0
    firstExcelEmailCol = 0

    secondExcelArray = []
    secondExcelFirstNameCol = 0
    secondExcelLastNameCol = 0
    secondExcelEmailCol = 0

    resultAfterEmailCheck = []
    resultAfterNameCheck = []

    remaining1AfterEmailCheck = []
    remaining1AfterNameCheck = []

    remaining2AfterEmailCheck = []
    remaining2AfterNameCheck = []

    for cell in wsFirst[1]:
        if(cell.value == 'FirstName'):
            firstExcelFirstNameCol = cell.column
        if(cell.value == 'LastName'):
            firstExcelLastNameCol = cell.column
        if(cell.value == 'Email'):
            firstExcelEmailCol = cell.column

    for cell in wsSecond[1]:
        if(cell.value == 'FirstName'):
            secondExcelFirstNameCol = cell.column
        if(cell.value == 'LastName'):
            secondExcelLastNameCol = cell.column
        if(cell.value == 'Email'):
            secondExcelEmailCol = cell.column

    for i in range(2, wsFirst.max_row + 1):
        firstExcelArray.append(RowObject(wsFirst[i][int(firstExcelFirstNameCol) - 1].value, wsFirst[i][firstExcelLastNameCol - 1].value, wsFirst[i][firstExcelEmailCol - 1].value))

    for i in range(2, wsSecond.max_row + 1):
        secondExcelArray.append(RowObject(wsSecond[i][int(secondExcelFirstNameCol) - 1].value, wsSecond[i][int(secondExcelLastNameCol) - 1].value, wsSecond[i][int(secondExcelEmailCol) - 1].value))

    # Compare Function
    def compareEmails(firstExcelArray, secondExcelArray):
        result = []
        remaining1 = []
        remaining2 = []
        for i in range(0, len(firstExcelArray)):
            if any(firstExcelArray[i].email == x.email for x in secondExcelArray) and firstExcelArray[i].email != 'none':
                result.append(firstExcelArray[i])
            elif firstExcelArray[i].email != 'none':
                remaining1.append(firstExcelArray[i])
        
        for i in range(0, len(secondExcelArray)):
            if any(secondExcelArray[i].email == x.email for x in firstExcelArray) and secondExcelArray[i].email != 'none':
                pass
            elif secondExcelArray[i].email != 'none':
                remaining2.append(secondExcelArray[i])
        
        return [result, remaining1, remaining2]
    
    def compareNames(firstExcelArray, secondExcelArray):
        result = []
        remaining1 = []
        remaining2 = []
        for i in range(0, len(firstExcelArray)):
            if any((firstExcelArray[i].firstName == x.firstName and firstExcelArray[i].lastName == x.lastName) for x in secondExcelArray) and (firstExcelArray[i].firstName != 'none' and firstExcelArray[i].lastName != 'none'):
                # newMail = str(firstExcelArray[i].email) + str(x.email)
                # newExcelValue = RowObject(firstExcelArray[i].firstName, firstExcelArray[i].lastName, newMail)
                result.append(firstExcelArray[i])
            elif firstExcelArray[i].firstName != 'none' and firstExcelArray[i].lastName != 'none':
                remaining1.append(firstExcelArray[i])
        
        for i in range(0, len(secondExcelArray)):
            if any((secondExcelArray[i].firstName == x.firstName and secondExcelArray[i].lastName == x.lastName) for x in firstExcelArray) and (secondExcelArray[i].firstName != 'none' and secondExcelArray[i].lastName != 'none'):
                pass
            elif secondExcelArray[i].firstName != 'none' and secondExcelArray[i].lastName != 'none':
                remaining2.append(secondExcelArray[i])
        
        return [result, remaining1, remaining2]
    



    checkEmails= compareEmails(firstExcelArray, secondExcelArray)
    resultAfterEmailCheck = checkEmails[0]
    remaining1AfterEmailCheck = checkEmails[1]
    remaining2AfterEmailCheck = checkEmails[2]

    checkNames = compareNames(remaining1AfterEmailCheck, remaining2AfterEmailCheck)
    resultAfterNameCheck = checkNames[0]
    remaining1AfterNameCheck = checkNames[1]
    remaining2AfterNameCheck = checkNames[2]

    with open('results.txt', 'w') as f:
        f.write("Email Comparison:\n")
        for i in range(0, len(resultAfterEmailCheck)):
            f.write("FirstName: %s LastName: %s Email: %s \n" % (resultAfterEmailCheck[i].firstName, resultAfterEmailCheck[i].lastName, resultAfterEmailCheck[i].email))

        f.write("\n\n")
        f.write("Full Name Comparison:\n")
        for i in range(0, len(resultAfterNameCheck)):
            f.write("FirstName: %s LastName: %s Email: %s \n" % (resultAfterNameCheck[i].firstName, resultAfterNameCheck[i].lastName, resultAfterNameCheck[i].email))
        
        f.write("\n\n")
        f.write("First Excel Unmatched:\n")
        for i in range(0, len(remaining1AfterNameCheck)):
            f.write("FirstName: %s LastName: %s Email: %s \n" % (remaining1AfterNameCheck[i].firstName, remaining1AfterNameCheck[i].lastName, remaining1AfterNameCheck[i].email))
        
        f.write("\n\n")
        f.write("Second Excel Unmatched:\n")
        for i in range(0, len(remaining2AfterNameCheck)):
            f.write("FirstName: %s LastName: %s Email: %s \n" % (remaining2AfterNameCheck[i].firstName, remaining2AfterNameCheck[i].lastName, remaining2AfterNameCheck[i].email))


selectFilesFunction()
# Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
# filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
# print(filename)

# tkFileDialog.askopenfilename()

# root = Tkinter.Tk()
# root.withdraw()



# outputFile = create an empty output file

# while recording, always use toLower and removeSpaces function to remove any unwanted strings

# firstMails = Match the column names for Email with itself for 1st

# secondMails = Match the column names for Email with itself for 2nd

# Iterate through 1st check if 2nd has the same email looking for CN+E
# If so, remove the entry from 1st and 2nd and add relative info to output

#     for 1st
#         for 2nd
#             if matches
#                 addToOutput()
#                 remove from 1st
#                 remove from 2nd

# Now you have the remaining files that matched all of the emails

# Further, you need to do the same for FullNames

# firstExcelNames = Match the column names for FirstName and LastName with themselves for 1st
#     Create an array with the values and fill that array by iterating through 1st

# secondExcelNames = Match the column names for FirstName and LastName with themselves for 2nd
#     Create an array with the values and fill that array by iterating through 2nd


# Iterate through 1st check if 2nd has the same person looking for CN+FN+LN
# If so, remove the entry from 1st and 2nd and add relative info to output
#     for 1st
#         for 2nd
#             if matches
#                 addToOutput()
#                 remove from 1st
#                 remove from 2nd


# Then add the remaining to the output as not related